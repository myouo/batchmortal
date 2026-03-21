import csv
import os
import re

import openpyxl

CSV_COLUMNS = [
    "nickname",
    "mode",
    "uuid",
    "paipuUrl",
    "startTime",
    "endTime",
    "resultUrl",
    "modelTag",
    "rating",
    "aiConsistencyRate",
    "aiConsistencyNumerator",
    "aiConsistencyDenominator",
    "temperature",
    "gameLength",
    "playerId",
    "reviewDuration",
    "screenshotPath",
    "timestamp",
]


def parse_metadata(metadata: dict) -> dict:
    """
    Parse the metadata dict returned by browser.py into typed fields.
    """

    def get(keys: list[str]) -> str:
        for raw_key, value in metadata.items():
            key = str(raw_key)
            lower_key = key.lower()
            for candidate in keys:
                if candidate.lower() in lower_key:
                    return value
        return ""

    ai_consistency = get(["\u4e00\u81f4\u7387", "Match Rate"])
    numerator, denominator, rate = "", "", ""

    # Matches strings like "195/271 = 71.956%".
    match = re.search(r"(\d+)\s*/\s*(\d+)\s*=\s*([\d.]+)%", ai_consistency)
    if match:
        numerator = match.group(1)
        denominator = match.group(2)
        rate = match.group(3) + "%"

    return {
        "modelTag": get(["model tag"]),
        "rating": get(["rating"]),
        "aiConsistencyRate": rate,
        "aiConsistencyNumerator": numerator,
        "aiConsistencyDenominator": denominator,
        "temperature": get(["temperature", "\u6e29\u5ea6"]),
        "gameLength": get(["\u5bf9\u5c40\u957f\u5ea6", "length"]),
        "playerId": get(["\u73a9\u5bb6 ID", "player"]),
        "reviewDuration": get(["\u5ba1\u67e5\u7528\u65f6", "Duration"]),
    }


class ResultWriter:
    """
    Keep the output file open and flush in batches to avoid O(n^2) XLSX writes.
    """

    def __init__(self, filepath: str, output_format: str = "csv", flush_every: int = 20):
        self.filepath = filepath
        self.output_format = output_format
        self.flush_every = max(1, flush_every)
        self.is_new = not os.path.exists(filepath)
        self._pending_rows = 0
        self._file = None
        self._csv_writer = None
        self._workbook = None
        self._worksheet = None

        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        if output_format == "csv":
            self._file = open(filepath, mode="a", newline="", encoding="utf-8")
            self._csv_writer = csv.writer(self._file)
            if self.is_new:
                self._csv_writer.writerow(CSV_COLUMNS)
                self._file.flush()
        elif output_format == "xlsx":
            if self.is_new:
                self._workbook = openpyxl.Workbook()
                self._worksheet = self._workbook.active
                self._worksheet.append(CSV_COLUMNS)
                self._uuid_to_row = {}
            else:
                self._workbook = openpyxl.load_workbook(filepath)
                self._worksheet = self._workbook.active
                self._uuid_to_row = {}
                headers = []
                for idx, row in enumerate(self._worksheet.iter_rows(values_only=True), start=1):
                    if idx == 1:
                        headers = [str(c) if c else "" for c in row]
                    else:
                        if "uuid" in headers:
                            uuid_idx = headers.index("uuid")
                            if len(row) > uuid_idx and row[uuid_idx]:
                                self._uuid_to_row[str(row[uuid_idx]).strip()] = idx
        else:
            raise ValueError(f"Unsupported output format: {output_format}")

    def write_row(self, row: dict):
        safe_row = [row.get(column, "") for column in CSV_COLUMNS]

        if self.output_format == "csv":
            self._csv_writer.writerow(safe_row)
            self._pending_rows += 1
            if self._pending_rows >= self.flush_every:
                self.flush()
            return

        uuid_val = str(row.get("uuid", "")).strip()
        if uuid_val and hasattr(self, "_uuid_to_row") and uuid_val in self._uuid_to_row:
            row_idx = self._uuid_to_row[uuid_val]
            for col_idx, val in enumerate(safe_row, start=1):
                self._worksheet.cell(row=row_idx, column=col_idx, value=val)
        else:
            self._worksheet.append(safe_row)
            if uuid_val and hasattr(self, "_uuid_to_row"):
                self._uuid_to_row[uuid_val] = self._worksheet.max_row

        self._pending_rows += 1
        if self._pending_rows >= self.flush_every:
            self.flush()

    def flush(self):
        if self._pending_rows == 0:
            return

        if self.output_format == "csv":
            self._file.flush()
        else:
            self._workbook.save(self.filepath)

        self._pending_rows = 0

    def close(self):
        try:
            self.flush()
        finally:
            if self._workbook is not None:
                self._workbook.close()
                self._workbook = None
                self._worksheet = None
            if self._file is not None:
                self._file.close()
                self._file = None
                self._csv_writer = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()


def append_row(filepath: str, row: dict, output_format: str = "csv"):
    """
    Backward-compatible one-shot append helper.
    """
    with ResultWriter(filepath, output_format=output_format, flush_every=1) as writer:
        writer.write_row(row)


def get_processed_uuids(filepath: str, output_format: str = "xlsx") -> set[str]:
    """
    Reads the existing output file and returns a set of all processed UUIDs.
    """
    processed = set()
    if not os.path.exists(filepath):
        return processed

    try:
        if output_format == "csv":
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if "uuid" in row and row["uuid"]:
                        if row.get("rating", "") != "ERROR":
                            processed.add(row["uuid"])
        elif output_format == "xlsx":
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            rows = iter(ws.rows)
            first_row = next(rows, None)
            if first_row is not None:
                headers = [cell.value for cell in first_row]
                if "uuid" in headers:
                    uuid_idx = headers.index("uuid")
                    rating_idx = headers.index("rating") if "rating" in headers else -1
                    for row in rows:
                        if len(row) > uuid_idx and row[uuid_idx].value:
                            rating_val = str(row[rating_idx].value).strip() if rating_idx >= 0 and len(row) > rating_idx and row[rating_idx].value is not None else ""
                            if rating_val != "ERROR":
                                processed.add(str(row[uuid_idx].value).strip())
            wb.close()
    except Exception as e:
        print(f"Failed to read processed UUIDs from {filepath}: {e}")
        
    return processed
